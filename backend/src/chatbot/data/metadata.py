"""Chargement des fichiers de référence (dictionnaire + métadonnées par logement).

Trois des quatre fichiers source décrivent le parc, pas les séries temporelles :

- ``Dictionnaire des données.xlsx`` : pour chaque canal brut (``libellé``), sa
  grandeur physique et son unité. Sert à annoter les variables du Zarr.
- ``Métadonnées.xlsx`` : matrice ``attribut × 100 logements`` (caractéristiques
  statiques : type de source froide air/eau vs géothermie, SCOP déclaré, etc.).
- ``Description des données brutes.xlsx`` : matrice ``attribut × 100 logements``
  décrivant la configuration d'installation et la disponibilité capteur (dont le
  drapeau « appoint joule inclus dans la mesure thermique »).

Les deux matrices sont fusionnées en une table indexée par identifiant de
logement (ex. ``"002026"``), qui devient la dimension ``logement`` du Dataset.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass

import openpyxl
import pandas as pd

from chatbot import paths

# Valeurs représentant une absence d'information dans les fichiers source.
_MISSING = {None, "", "-", "—", "n/a", "na", "none"}


def slugify(label: str) -> str:
    """Transforme un libellé humain en identifiant snake_case sans accents."""
    text = unicodedata.normalize("NFKD", str(label))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _dedup(names: list[str]) -> list[str]:
    """Suffixe les collisions de slugs (``x``, ``x_2``, ``x_3``...)."""
    seen: dict[str, int] = {}
    out: list[str] = []
    for name in names:
        if name in seen:
            seen[name] += 1
            out.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 1
            out.append(name)
    return out


def _logement_id(header: str) -> str:
    """``"Log. 001026"`` -> ``"001026"``. Zéro-pad à 6 (Excel mange les zéros de tête)."""
    s = str(header).replace("Log.", "").strip()
    return s.zfill(6) if s.isdigit() else s


def _coerce(value: object) -> object:
    """Convertit une cellule : nombre si possible, sinon chaîne, sinon ``None``."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text.lower() in _MISSING:
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return text


@dataclass(frozen=True)
class ChannelMeta:
    """Métadonnée d'un canal brut, issue du dictionnaire."""

    name: str
    quantity: str
    unit: str


def load_dictionary() -> dict[str, ChannelMeta]:
    """Mappe ``libellé -> ChannelMeta`` depuis le dictionnaire des données."""
    wb = openpyxl.load_workbook(paths.DICTIONARY_XLSX, read_only=True, data_only=True)
    rows = list(wb.worksheets[0].iter_rows(values_only=True))
    out: dict[str, ChannelMeta] = {}
    # Ligne 1 = en-têtes ; colonnes : [_, libellé, Nom, Grandeur, Unité].
    for row in rows[2:]:
        label = row[1]
        if label is None:
            continue
        out[str(label).strip()] = ChannelMeta(
            name=str(row[2] or "").strip(),
            quantity=str(row[3] or "").strip(),
            unit=str(row[4] or "").strip(),
        )
    return out


def _load_matrix(path, label_col: int, first_log_col: int) -> pd.DataFrame:
    """Lit une matrice ``attribut × logements`` -> DataFrame (index = logement).

    ``label_col`` : index 0-based de la colonne portant le nom de l'attribut.
    ``first_log_col`` : index 0-based de la première colonne ``Log. NNNNNN``.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = list(wb.worksheets[0].iter_rows(values_only=True))
    header = rows[1]
    log_ids = [_logement_id(c) for c in header[first_log_col:] if c]

    attr_names: list[str] = []
    columns: dict[str, list[object]] = {}
    for row in rows[2:]:
        label = row[label_col]
        if label is None:
            continue
        attr_names.append(slugify(label))
        values = [_coerce(v) for v in row[first_log_col : first_log_col + len(log_ids)]]
        columns[attr_names[-1]] = values

    slugs = _dedup(attr_names)
    data = {slug: columns[orig] for slug, orig in zip(slugs, attr_names)}
    return pd.DataFrame(data, index=pd.Index(log_ids, name="logement"))


def load_fleet_metadata() -> pd.DataFrame:
    """Fusionne les deux matrices en une table statique des 100 logements.

    L'index est l'identifiant de logement (ex. ``"002026"``) ; chaque colonne est
    un attribut statique (slug). Préfixe ``inst_`` pour les attributs issus de la
    description des données brutes afin d'éviter les collisions et de tracer la
    source.
    """
    meta = _load_matrix(paths.METADATA_XLSX, label_col=1, first_log_col=4)
    desc = _load_matrix(paths.DESCRIPTION_XLSX, label_col=2, first_log_col=5)
    desc = desc.add_prefix("inst_")
    fleet = meta.join(desc, how="outer")
    return fleet
